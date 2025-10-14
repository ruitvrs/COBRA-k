"""Functions for generating spreadsheet overviews of variability/optimization results"""

# IMPORT SECTION #
from dataclasses import dataclass, field
from math import exp, log
from statistics import mean, median

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ConfigDict, validate_call

from .constants import (
    ALPHA_VAR_PREFIX,
    DF_VAR_PREFIX,
    ENZYME_VAR_INFIX,
    ENZYME_VAR_PREFIX,
    ERROR_VAR_PREFIX,
    GAMMA_VAR_PREFIX,
    IOTA_VAR_PREFIX,
    KAPPA_VAR_PREFIX,
    LNCONC_VAR_PREFIX,
    OBJECTIVE_VAR_NAME,
    PROT_POOL_REAC_NAME,
    SOLVER_STATUS_KEY,
    TERMINATION_CONDITION_KEY,
)
from .dataclasses import Enzyme, Metabolite, Model, Reaction
from .utilities import (
    compare_multiple_results_to_best,
    get_df_and_efficiency_factors_sorted_lists,
    get_full_enzyme_mw,
    get_fwd_rev_corrected_flux,
    get_metabolite_consumption_and_production,
    get_reaction_enzyme_var_id,
    get_reaction_string,
    get_unoptimized_reactions_in_nlp_solution,
)

# CONSTANTS SECTION #
WIDTH_DEFAULT = 12
"""Default spreadsheel column width"""
ABS_EPSILON = 1e-12
"""Lower absolute values are shown as 0 in the spreadsheet"""

FONT_DEFAULT = Font(name="Calibri")
"""Default font for spreadsheet cells"""
FONT_BOLD = Font(name="Calibri", bold=True)
"""Bold font for spreadsheet cells"""
FONT_BOLD_AND_UNDERLINED = Font(name="Calibri", bold=True, underline="single")
"""Bold and underlined font for spreadsheet cells"""
FONT_ITALIC = Font(
    name="Calibri",
    italic=True,
)
FONT_BLACK = Font(name="Calibri", color="000000")

BG_COLOR_DEFAULT = PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
)
BG_COLOR_GREEN = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)
BG_COLOR_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
BG_COLOR_BLUE = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
BG_COLOR_BLACK = PatternFill(
    start_color="000000", end_color="000000", fill_type="solid"
)

SIDE_LEFT_BORDER = Side(
    border_style="medium", color="000000"
)  # "thin"/"medium"/"thick"/... can be set
BORDER_BLACK_LEFT = Border(left=SIDE_LEFT_BORDER)
BORDER_DEFAULT = Border()


SCALE_THREE_COLORS = ColorScaleRule(
    start_type="min",
    start_color="00FF00",
    mid_type="percentile",
    mid_value=50,
    mid_color="FFFF00",
    end_type="max",
    end_color="FF0000",
)
SCALE_TWO_COLORS = ColorScaleRule(
    start_type="min", start_color="00FF00", end_type="max", end_color="FFFF00"
)


# DATACLASSES SECTION #
@dataclass
class OptimizationDataset:
    """Represents an optimization result and which of its data shall be shown in the spreadsheet"""

    data: dict[str, float]
    """The optimization result"""
    with_df: bool = False
    """Shall driving forces be shown in the spreadsheet?"""
    with_vplus: bool = False
    """Shall V+ values be shown in the spreadsheet?"""
    with_kappa: bool = False
    """Shall saturation term values be shown in the spreadsheet?"""
    with_gamma: bool = False
    """Shall gamma values be shown in the spreadsheet?"""
    with_iota: bool = False
    """Shall iota values (inhibition terms) be shown in the spreadsheet?"""
    with_alpha: bool = False
    """Shall alpha values (activation terms) be shown in the spreadsheet?"""
    with_kinetic_differences: bool = False
    """Shall differences between NLP fluxes and 'real' fluxes from kinetics be shown in the spreadsheet?"""
    with_error_corrections: bool = False
    """Shall error corrections be shown as their own sheet?"""


@dataclass
class SpreadsheetCell:
    """Represents the content of a spreadsheet cell.

    Includes the shown value, background color, font style
    and border setting.
    """

    value: float | str | int | bool | None
    """The cell's shown content value (if None, nothing is shown)"""
    bg_color: PatternFill = field(default=BG_COLOR_DEFAULT)
    """The cell's background color (default: BG_COLOR_DEFAULT)"""
    font: Font = field(default=FONT_DEFAULT)
    """The cell's font style (default: FONT_DEFAULT)"""
    border: Border | None = field(default=None)
    """The cell's border style (None if no style given; default: None)"""


EMPTY_CELL = SpreadsheetCell(None)
"""Represents an empty spreadsheeet cell without content"""


@dataclass
class Title:
    """Represents a title or metatitle used in visualizations."""

    text: str
    """Title text content"""
    width: float
    """With of column"""
    is_metatitle: bool = field(default=False)
    """If True, the title is shown *under* a the major title line in a second line. Defaults to False."""


@dataclass
class VariabilityDataset:
    """Represents a dataset with variability for plotting, including error bars or ranges."""

    data: dict[str, tuple[float, float]]
    """The variability data dict, as returned by COBRAk's variability functions"""
    with_df: bool = False
    """Shall driving force variabilities be shown?"""


# "PRIVATE" FUNCTIONS SECTION #
@validate_call()
def sum_concs(
    result: dict[str, float],
    conc_sum_include_suffixes: list[str],
    conc_sum_ignore_prefixes: list[str],
) -> float:
    """Returns the exponentiated concentration of all metabolites in  the result."""
    concsum = 0.0
    for key, value in result.items():
        if key.startswith(LNCONC_VAR_PREFIX):
            met_id = key[len(LNCONC_VAR_PREFIX) :]
            if any(met_id.startswith(prefix) for prefix in conc_sum_ignore_prefixes):
                continue
            if not any(met_id.endswith(suffix) for suffix in conc_sum_include_suffixes):
                continue
            concsum += exp(value)
    return concsum


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def _create_xlsx_from_datadicts(
    path: str,
    titles_and_data_dict: dict[
        str,
        tuple[
            list[Title],
            dict[str, list[str | float | int | bool | None | SpreadsheetCell]],
        ],
    ],
) -> None:
    """Create and save an xlsx file from a dictionary of data and titles.

    This function generates an Excel workbook from the provided data dictionary, where each key represents a sheet name and
    its value is a tuple containing a list of titles and a data dictionary.
    The titles are used as headers, and the data dictionary contains the rows of data for each sheet.

    Args:
        path (str): The file path where the Excel workbook will be saved.
        titles_and_data_dict (dict): A dictionary where each key is a sheet name and its value is a tuple containing:
            - A list of Title objects representing the column headers.
            - A dictionary mapping item IDs to lists of cell values (which can be of various types including SpreadsheetCell).
    """
    wb = Workbook()

    for sheet_name, titles_and_data in titles_and_data_dict.items():
        wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        titles = titles_and_data[0]
        has_metatitles = bool(sum(title.is_metatitle for title in titles))
        current_column = 1

        # Set sheet titles
        for title in titles:
            cell = SpreadsheetCell(title.text)
            if title.is_metatitle:
                cell.font = FONT_ITALIC
                _set_cell(sheet, 1, current_column, cell)
                continue
            cell.font = FONT_BOLD
            line = 2 if has_metatitles else 1
            _set_cell(sheet, line, current_column, cell)
            current_column += 1

        # Freeze spreadsheet rows according to title height
        datadict = titles_and_data[1]
        if titles == []:
            start_line = 1
        elif not has_metatitles:
            start_line = 2
            sheet.freeze_panes = "B2"
        else:
            start_line = 3
            sheet.freeze_panes = "B3"

        current_line = start_line

        # Fill in the data
        if all(key.isdigit() for key in datadict):
            sorted_item_ids = sorted(datadict.keys(), key=lambda x: int(x))  # noqa: PLW0108
        else:
            sorted_item_ids = sorted(datadict.keys())
        for item_id in sorted_item_ids:
            datalist = datadict[item_id]
            current_column = 1
            for value in datalist:
                if isinstance(value, SpreadsheetCell):
                    cell = value
                else:
                    cell = SpreadsheetCell(value)  # type: ignore
                if cell.bg_color == BG_COLOR_BLACK:
                    cell.font = FONT_BLACK
                _set_cell(sheet, current_line, current_column, cell)
                current_column += 1
            current_line += 1

        num_columns = current_column

        # Apply conditional formatting for specific sheets
        if sheet_name == "Complexes":
            for current_column in range(num_columns):
                column_letter = get_column_letter(current_column + 1)
                range_string = f"{column_letter}{start_line}:{column_letter}10000"
                sheet.conditional_formatting.add(range_string, SCALE_TWO_COLORS)

    # Remove the default sheet created by openpyxl
    wb.remove(wb["Sheet"])
    wb.save(path)


@validate_call(config=ConfigDict(arbitrary_types_allowed=True), validate_return=True)
def _get_empty_cell() -> SpreadsheetCell:
    """Returns spreadsheet cell with no content in full black"""
    return SpreadsheetCell(None, bg_color=BG_COLOR_BLACK, font=FONT_BLACK)


@validate_call(validate_return=True)
def _get_enzcomplex_reaction(
    cobrak_model: Model, enzcomplex_id: str
) -> tuple[str, Reaction]:
    """Retrieve the reaction associated with a given enzyme complex ID in a COBRAk model.

    Args:
        cobrak_model (Model): The COBRAk model containing reactions.
        enzcomplex_id (str): The enzyme complex ID from which to extract the reaction ID.

    Returns:
        tuple[str, Reaction]: A tuple containing the reaction ID and the corresponding Reaction object.
    """
    reac_id = enzcomplex_id.split(ENZYME_VAR_INFIX)[1]
    return reac_id, cobrak_model.reactions[reac_id]


@validate_call(validate_return=True)
def _get_met_id_from_met_var_id(met_var_id: str) -> str:
    """Gives the (N/MI)LP metabolite concentration variable ID, derived from the metabolite ID"""
    return ("\b" + met_var_id).replace("\b" + LNCONC_VAR_PREFIX, "").replace("\b", "")


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def _get_optimization_bg_color(opt_value: float) -> PatternFill:
    """Determine the background color for a cell based on an optimization value.

    Args:
        opt_value (float): The optimization value to evaluate.

    Returns:
        PatternFill: The background color to be used for the cell.
    """
    if abs(opt_value) <= ABS_EPSILON:
        return BG_COLOR_BLACK
    return BG_COLOR_GREEN


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def _get_variability_bg_color(min_value: float, max_value: float) -> PatternFill:
    """Determine the background color for a cell based on variability values.

    Args:
        min_value (float): The minimum value to evaluate.
        max_value (float): The maximum value to evaluate.

    Returns:
        PatternFill: The background color to be used for the cell.
    """
    if min_value > ABS_EPSILON:
        return BG_COLOR_RED
    if abs(max_value) <= ABS_EPSILON:
        return BG_COLOR_BLACK
    return BG_COLOR_GREEN


@validate_call(validate_return=True)
def _na_str_or_value(
    value: str | float | int | bool | None,
) -> str | float | int | bool:
    """Returns the value if it is not None, otherwise, 'N/A'

    Args:
        value (str | float | int | bool | None): The value to check
    """
    if value is None:
        return " "
    return value


@validate_call(validate_return=True)
def _num_to_sheet_letter(number: int) -> str:
    """Convert a given column number to its corresponding spreadsheet column letter.

    Args:
        number (int): The column number to convert (1-based index).

    Returns:
        str: The corresponding spreadsheet column letter.

    Example:
        >>> _num_to_sheet_letter(1)
        'A'
        >>> _num_to_sheet_letter(28)
        'AB'
    """
    column = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        column = chr(65 + remainder) + column
    return column


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def _set_cell(
    sheet: Worksheet,
    line: int,
    column: int,
    cell: SpreadsheetCell,
) -> None:
    """Set the properties of a cell.

    Updates the value, background color, font, and border of a specified cell in a spreadsheet.

    Args:
        sheet: The openpyxl worksheet object where the cell is located.
        line (int): The row number of the cell (1-based index).
        column (int): The column number of the cell (1-based index).
        cell (SpreadsheetCell): An object containing the properties to be set for the cell, including value, background color, font, and border.
    """
    openpyxl_cell = sheet.cell(row=line, column=column)
    openpyxl_cell.value = _na_str_or_value(cell.value)
    if cell.bg_color is not None:
        openpyxl_cell.fill = cell.bg_color
    if cell.font is not None:
        openpyxl_cell.font = cell.font
    if cell.border is not None:
        openpyxl_cell.border = cell.border


# "PUBLIC" FUNCTIONS SECTION #
@validate_call
def create_cobrak_spreadsheet(
    path: str,
    cobrak_model: Model,
    variability_datasets: dict[str, VariabilityDataset],
    optimization_datasets: dict[str, OptimizationDataset],
    is_maximization: bool = True,
    sheet_description: list[str] = [],
    min_var_value: float = 1e-6,
    min_rel_correction: float = 0.01,
    kinetic_difference_precision: int = 6,
    objective_overwrite: None | str = None,
    extra_optstatistics_data: dict[str, list[str | float | int | bool | None]] = {},
    show_regulation_coefficients: bool = True,
) -> None:
    """Generates a comprehensive Excel spreadsheet summarizing variability and optimization results for a COBRAk model.

    This function creates an Excel file that organizes and visualizes various aspects of the model's reactions, metabolites, enzymes, and optimization results.
    It includes multiple sheets, each focusing on different components of the model and their corresponding data.

    In particular, the generated Excel workbook includes the following sheets:

    1. **Index**: Provides an overview of the different sections in the spreadsheet.
    2. **A) Optimization statistics**: Displays statistical summaries of the optimization results, including objective values, solver status, and flux comparisons.
    3. **B) Model settings**: Lists the model's parameters such as protein pool, gas constant, temperature, and annotations.
    4. **C) Reactions**: Details each reaction's properties, including reaction strings, ΔG'° values, enzyme associations, and kinetic parameters.
    5. **D) Metabolites**: Shows metabolite concentrations, their ranges, and annotations.
    6. **E) Enzymes**: Lists individual enzymes with their molecular weights and concentration ranges.
    7. **F) Complexes**: Provides information on enzyme complexes, including associated reactions and molecular weights.
    8. **G) Corrections (optional)**: If error corrections are included in the optimization datasets, this sheet displays the corrections applied.

    Each sheet is populated with data from the provided variability and optimization datasets, formatted for readability with appropriate styling,
    including background colors and borders to highlight important information.

    The function also handles various edge cases, such as missing data and low-flux reactions, ensuring that the spreadsheet remains organized and informative.

    Args:
        path (str): The file path where the Excel workbook will be saved.
        cobrak_model (Model): The COBRAk model containing reactions, metabolites, and enzymes.
        variability_datasets (dict[str, VariabilityDataset]): A dictionary of variability datasets, where each key is a dataset name and the value contains
                                                             the data and flags for what to display.
        optimization_datasets (dict[str, OptimizationDataset]): A dictionary of optimization results, where each key is a dataset name and the value contains
                                                                the optimization data and flags for what to display.
        is_maximization (bool, optional): Indicates whether the optimization is a maximization problem. Defaults to True.
        sheet_description (list[str], optional): A list of description lines to include in the index sheet. Defaults to an empty list.
        min_var_value (float, optional): Where applicable (e.g. for fluxes), the minimum value to display a variable's value. Does not apply for error correction value (see next argument for that.
                                         Defaults to 1e-6.
        min_rel_correction (float, optional): Minimal relative change to associated original value for which an error correction value is shown.
        kinetic_difference_precision (int, optional): The number of decimal places to round kinetic differences. Defaults to 6.

    Returns:
        None: The function does not return any value but saves the Excel workbook to the specified path.
    """
    all_reac_ids = list(cobrak_model.reactions.keys())
    all_met_ids = list(cobrak_model.metabolites.keys())
    all_enzyme_ids = list(cobrak_model.enzymes.keys())
    all_met_var_ids = [LNCONC_VAR_PREFIX + met_id for met_id in all_met_ids]
    all_enzcomplex_ids = []
    for reac_id, reaction in cobrak_model.reactions.items():
        if reaction.enzyme_reaction_data is None:
            continue
        all_enzcomplex_ids.append(get_reaction_enzyme_var_id(reac_id, reaction))

    has_any_vplus = any(
        opt_data.with_vplus for opt_data in optimization_datasets.values()
    )
    has_any_df = any(opt_data.with_df for opt_data in optimization_datasets.values())
    has_any_kappa = any(
        opt_data.with_kappa for opt_data in optimization_datasets.values()
    )
    has_any_gamma = any(
        opt_data.with_gamma for opt_data in optimization_datasets.values()
    )
    has_any_iota = any(
        opt_data.with_iota for opt_data in optimization_datasets.values()
    )
    has_any_alpha = any(
        opt_data.with_alpha for opt_data in optimization_datasets.values()
    )
    has_any_kinetic_differences = any(
        opt_data.with_kinetic_differences for opt_data in optimization_datasets.values()
    )

    kappa_gamma_iota_alpha_str_list = []
    if has_any_kappa:
        kappa_gamma_iota_alpha_str_list.append("κ")
    if has_any_gamma:
        kappa_gamma_iota_alpha_str_list.append("γ")
    if has_any_iota:
        kappa_gamma_iota_alpha_str_list.append("ι")
    if has_any_alpha:
        kappa_gamma_iota_alpha_str_list.append("α")
    kappa_gamma_iota_alpha_str = "⋅".join(kappa_gamma_iota_alpha_str_list)

    # Index sheet
    index_titles: list[Title] = []
    index_cells: dict[str, list[str | float | int | bool | None | SpreadsheetCell]] = {}
    sheet_line = 1
    for description_line in sheet_description:
        index_cells[_num_to_sheet_letter(sheet_line)] = [
            SpreadsheetCell(
                description_line,
            ),
        ]
        sheet_line += 1

    if sheet_line == 1:
        sheet_line = 0  # No description lines provided, first line can be used for A) to ... as following
    index_cells = {
        _num_to_sheet_letter(sheet_line + 1): [
            SpreadsheetCell(
                "A) Optimization statistics: Objective values, minimal/maximal occurring kinetic values, ...",
            ),
        ],
        _num_to_sheet_letter(sheet_line + 2): [
            SpreadsheetCell(
                "B) Global setting: Model settings such as the temperature, protein pool, ...",
            ),
        ],
        _num_to_sheet_letter(sheet_line + 3): [
            SpreadsheetCell(
                "C) Reactions: Their fluxes, driving forces, kinetic values...",
            ),
        ],
        _num_to_sheet_letter(sheet_line + 4): [
            SpreadsheetCell(
                "D) Metabolites: Their concentrations, formulas, ...",
            ),
        ],
        _num_to_sheet_letter(sheet_line + 5): [
            SpreadsheetCell(
                "E) Enzymes: The single enzymes occurring in the model with their concentration settings (if any given)",
            ),
        ],
        _num_to_sheet_letter(sheet_line + 6): [
            SpreadsheetCell(
                "F) Complexes: The (multi- or single-)enzyme complexes occurring in the model with protein pool fraction data",
            ),
        ],
    }

    if has_any_kappa or has_any_gamma or has_any_iota or has_any_alpha:
        index_cells |= {
            _num_to_sheet_letter(sheet_line + 6): [
                SpreadsheetCell(
                    "F) Complexes: The (multi- or single-)enzyme complexes occurring in the model with protein pool fraction data",
                ),
            ],
        }

    # Model settings sheet
    model_titles: list[Title] = []
    model_cells: dict[str, list[str | float | int | bool | None | SpreadsheetCell]] = {
        "A": [
            SpreadsheetCell("Protein pool [g⋅gDW⁻¹]", font=FONT_BOLD),
            SpreadsheetCell(cobrak_model.max_prot_pool),
        ],
        "B": [
            SpreadsheetCell("R [kJ⋅K⁻¹⋅mol⁻¹)]", font=FONT_BOLD),
            SpreadsheetCell(cobrak_model.R),
        ],
        "C": [
            SpreadsheetCell("T [K]", font=FONT_BOLD),
            SpreadsheetCell(cobrak_model.T),
        ],
        "D": [
            SpreadsheetCell("R⋅T [kJ⋅mol⁻¹]", font=FONT_BOLD),
            SpreadsheetCell(cobrak_model.R * cobrak_model.T),
        ],
        "E": [
            SpreadsheetCell("κ-ignored metabolites", font=FONT_BOLD),
            SpreadsheetCell(str(cobrak_model.kinetic_ignored_metabolites)),
        ],
        "F": [
            SpreadsheetCell("Model annotation", font=FONT_BOLD),
            SpreadsheetCell(str(cobrak_model.annotation)),
        ],
        "G": [
            SpreadsheetCell("Maximal concentration sum [M]:", font=FONT_BOLD),
            SpreadsheetCell(str(cobrak_model.max_conc_sum)),
        ],
        "H": [
            SpreadsheetCell("Metabolite pool [M]", font=FONT_BOLD),
            SpreadsheetCell(cobrak_model.max_conc_sum),
        ],
        "I": [
            SpreadsheetCell(
                "Metabolite pool ignore prefixes (i.e. metabolites with this prefix are not counted)",
                font=FONT_BOLD,
            ),
            SpreadsheetCell("; ".join(cobrak_model.conc_sum_ignore_prefixes)),
        ],
        "J": [
            SpreadsheetCell(
                "Metabolite pool include prefixes (i.e. only metabolites with this prefix are counted)",
                font=FONT_BOLD,
            ),
            SpreadsheetCell("; ".join(cobrak_model.conc_sum_include_suffixes)),
        ],
    }

    # Statistics sheet
    comparisons = compare_multiple_results_to_best(
        cobrak_model,
        [dataset.data for dataset in optimization_datasets.values()],
        is_maximization,
        min_var_value,
    )

    stats_titles: list[Title] = [Title("", WIDTH_DEFAULT)]
    stats_cells: dict[str, list[str | float | int | bool | None | SpreadsheetCell]] = {
        "0": [
            SpreadsheetCell(
                "Objective value"
                if objective_overwrite is None
                else f"{objective_overwrite} value",
                font=FONT_BOLD,
            ),
        ],
        "1": [
            SpreadsheetCell("Solver status (see COBRAk documentation)", font=FONT_BOLD),
        ],
        "2": [
            SpreadsheetCell(
                "Termination condition (see COBRAk documentaiton)", font=FONT_BOLD
            ),
        ],
    }

    statline = 3
    if has_any_vplus:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("Used protein pool [g⋅gDW⁻¹]", font=FONT_BOLD),
            ],
        }
        statline += 1

    if has_any_df or has_any_gamma:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell(
                    "Used metabolite concentration pool [M]", font=FONT_BOLD
                ),
            ],
        }
        statline += 1

    if has_any_df:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("Min driving force [kJ⋅mol⁻¹]", font=FONT_BOLD),
            ],
            f"{statline + 1}": [
                SpreadsheetCell("Max driving force [kJ⋅mol⁻¹]", font=FONT_BOLD),
            ],
            f"{statline + 2}": [
                SpreadsheetCell("Mean driving force [kJ⋅mol⁻¹]", font=FONT_BOLD),
            ],
            f"{statline + 3}": [
                SpreadsheetCell("Median driving force [kJ⋅mol⁻¹]", font=FONT_BOLD),
            ],
        }
        statline += 4

    if has_any_gamma:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("Min γ", font=FONT_BOLD),
            ],
            f"{statline + 1}": [
                SpreadsheetCell("Max γ", font=FONT_BOLD),
            ],
            f"{statline + 2}": [
                SpreadsheetCell("Mean γ", font=FONT_BOLD),
            ],
            f"{statline + 3}": [
                SpreadsheetCell("Median γ", font=FONT_BOLD),
            ],
        }
        statline += 4

    if has_any_kappa:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("Min κ", font=FONT_BOLD),
            ],
            f"{statline + 1}": [
                SpreadsheetCell("Max κ", font=FONT_BOLD),
            ],
            f"{statline + 2}": [
                SpreadsheetCell("Mean κ", font=FONT_BOLD),
            ],
            f"{statline + 3}": [
                SpreadsheetCell("Median κ", font=FONT_BOLD),
            ],
        }
        statline += 4

    if has_any_iota:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("Min ι", font=FONT_BOLD),
            ],
            f"{statline + 1}": [
                SpreadsheetCell("Max ι", font=FONT_BOLD),
            ],
            f"{statline + 2}": [
                SpreadsheetCell("Mean ι", font=FONT_BOLD),
            ],
            f"{statline + 3}": [
                SpreadsheetCell("Median ι", font=FONT_BOLD),
            ],
        }
        statline += 4

    if has_any_alpha:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("Min α", font=FONT_BOLD),
            ],
            f"{statline + 1}": [
                SpreadsheetCell("Max α", font=FONT_BOLD),
            ],
            f"{statline + 2}": [
                SpreadsheetCell("Mean α", font=FONT_BOLD),
            ],
            f"{statline + 3}": [
                SpreadsheetCell("Median α", font=FONT_BOLD),
            ],
        }
        statline += 4

    if has_any_kappa or has_any_gamma or has_any_alpha or has_any_iota:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell(f"Min {kappa_gamma_iota_alpha_str}", font=FONT_BOLD),
            ],
            f"{statline + 1}": [
                SpreadsheetCell(f"Max {kappa_gamma_iota_alpha_str}", font=FONT_BOLD),
            ],
            f"{statline + 2}": [
                SpreadsheetCell(f"Mean {kappa_gamma_iota_alpha_str}", font=FONT_BOLD),
            ],
            f"{statline + 3}": [
                SpreadsheetCell(f"Median {kappa_gamma_iota_alpha_str}", font=FONT_BOLD),
            ],
        }
        statline += 4

    stats_cells |= {
        f"{statline}": [
            SpreadsheetCell("Min flux difference to best", font=FONT_BOLD),
        ],
        f"{statline + 1}": [
            SpreadsheetCell("Max flux difference to best", font=FONT_BOLD),
        ],
        f"{statline + 2}": [
            SpreadsheetCell("Sum of flux differences to best", font=FONT_BOLD),
        ],
        f"{statline + 3}": [
            SpreadsheetCell("Mean flux difference to best", font=FONT_BOLD),
        ],
        f"{statline + 4}": [
            SpreadsheetCell("Median flux difference to best", font=FONT_BOLD),
        ],
        f"{statline + 5}": [
            SpreadsheetCell("Objective difference to best", font=FONT_BOLD),
        ],
        f"{statline + 6}": [
            SpreadsheetCell(
                "Only in this to best (regarding active reactions)", font=FONT_BOLD
            ),
        ],
        f"{statline + 7}": [
            SpreadsheetCell(
                "Only in best to this (regarding active reactions)", font=FONT_BOLD
            ),
        ],
    }
    statline += 8

    if has_any_kinetic_differences:
        stats_cells |= {
            f"{statline}": [
                SpreadsheetCell("'Really' used protein pool [g⋅gDW⁻¹]", font=FONT_BOLD),
            ]
        }
        statline += 1

    for extrai, extratitle in enumerate(extra_optstatistics_data.keys()):
        stats_cells[f"{statline + extrai}"] = [
            SpreadsheetCell(extratitle, font=FONT_BOLD)
        ]

    # Optimization data
    for current_dataset_i, (opt_dataset_name, opt_dataset) in enumerate(
        optimization_datasets.items()
    ):
        statline = 0
        stats_titles.append(Title(opt_dataset_name, WIDTH_DEFAULT))
        if objective_overwrite is None:
            stats_cells[f"{statline}"].append(opt_dataset.data[OBJECTIVE_VAR_NAME])
        else:
            stats_cells[f"{statline}"].append(opt_dataset.data[objective_overwrite])
        stats_cells[f"{statline + 1}"].append(opt_dataset.data[SOLVER_STATUS_KEY])
        stats_cells[f"{statline + 2}"].append(
            opt_dataset.data[TERMINATION_CONDITION_KEY]
        )
        statline += 3

        if has_any_vplus:
            if PROT_POOL_REAC_NAME in opt_dataset.data:
                stats_cells[f"{statline}"].append(opt_dataset.data[PROT_POOL_REAC_NAME])
                statline += 1
            else:
                stats_cells[f"{statline}"].append(_get_empty_cell())
                statline += 1

        if has_any_df or has_any_gamma:
            if any(x.startswith(LNCONC_VAR_PREFIX) for x in opt_dataset.data):
                stats_cells[f"{statline}"].append(
                    sum_concs(
                        opt_dataset.data,
                        cobrak_model.conc_sum_include_suffixes,
                        cobrak_model.conc_sum_ignore_prefixes,
                    )
                )
                statline += 1
            else:
                stats_cells[f"{statline}"].append(_get_empty_cell())
                statline += 1

        if opt_dataset.with_df:
            df_stats, _, _, _, _, _ = get_df_and_efficiency_factors_sorted_lists(
                cobrak_model,
                opt_dataset.data,
                min_var_value,
            )
            stats_cells[f"{statline}"].append(min(df_stats.values()))
            stats_cells[f"{statline + 1}"].append(max(df_stats.values()))
            stats_cells[f"{statline + 2}"].append(mean(df_stats.values()))
            stats_cells[f"{statline + 3}"].append(median(df_stats.values()))
            statline += 4
        elif has_any_df:
            for line_letter in (f"{statline + j}" for j in range(4)):
                stats_cells[line_letter].append(_get_empty_cell())
            statline += 4

        if opt_dataset.with_gamma:
            _, _, gamma_stats, _, _, _ = get_df_and_efficiency_factors_sorted_lists(
                cobrak_model,
                opt_dataset.data,
                min_var_value,
            )
            stats_cells[f"{statline}"].append(min(gamma_stats.values()))
            stats_cells[f"{statline + 1}"].append(max(gamma_stats.values()))
            stats_cells[f"{statline + 2}"].append(mean(gamma_stats.values()))
            stats_cells[f"{statline + 3}"].append(median(gamma_stats.values()))
            statline += 4
        elif has_any_gamma:
            for line_letter in (f"{statline + j}" for j in range(4)):
                stats_cells[line_letter].append(_get_empty_cell())
            statline += 4

        if opt_dataset.with_kappa:
            _, kappa_stats, _, _, _, _ = get_df_and_efficiency_factors_sorted_lists(
                cobrak_model,
                opt_dataset.data,
                min_var_value,
            )
            stats_cells[f"{statline}"].append(min(kappa_stats.values()))
            stats_cells[f"{statline + 1}"].append(max(kappa_stats.values()))
            stats_cells[f"{statline + 2}"].append(mean(kappa_stats.values()))
            stats_cells[f"{statline + 3}"].append(median(kappa_stats.values()))
            statline += 4
        elif has_any_kappa:
            for line_letter in (f"{statline + j}" for j in range(4)):
                stats_cells[line_letter].append(_get_empty_cell())
            statline += 4

        if opt_dataset.with_iota:
            iota_values = [
                opt_dataset.data[x]
                for x in opt_dataset.data
                if x.startswith(IOTA_VAR_PREFIX)
                and (opt_dataset.data[x[len(IOTA_VAR_PREFIX) :]] > min_var_value)
            ]
            stats_cells[f"{statline}"].append(min(iota_values))
            stats_cells[f"{statline + 1}"].append(max(iota_values))
            stats_cells[f"{statline + 2}"].append(mean(iota_values))
            stats_cells[f"{statline + 3}"].append(median(iota_values))
            statline += 4
        elif has_any_iota:
            for line_letter in (f"{statline + j}" for j in range(4)):
                stats_cells[line_letter].append(_get_empty_cell())
            statline += 4

        if opt_dataset.with_alpha:
            alpha_values = [
                opt_dataset.data[x]
                for x in opt_dataset.data
                if x.startswith(ALPHA_VAR_PREFIX)
                and (opt_dataset.data[x[len(ALPHA_VAR_PREFIX) :]] > min_var_value)
            ]
            stats_cells[f"{statline}"].append(min(alpha_values))
            stats_cells[f"{statline + 1}"].append(max(alpha_values))
            stats_cells[f"{statline + 2}"].append(mean(alpha_values))
            stats_cells[f"{statline + 3}"].append(median(alpha_values))
            statline += 4
        elif has_any_alpha:
            for line_letter in (f"{statline + j}" for j in range(4)):
                stats_cells[line_letter].append(_get_empty_cell())
            statline += 4

        if (
            opt_dataset.with_kappa
            or opt_dataset.with_gamma
            or opt_dataset.with_alpha
            or opt_dataset.with_iota
        ):
            _, _, _, _, _, multiplier_stats = (
                get_df_and_efficiency_factors_sorted_lists(
                    cobrak_model,
                    opt_dataset.data,
                    min_var_value,
                )
            )
            efficiencies_product_stats_values = [
                x[0] for x in multiplier_stats.values()
            ]
            stats_cells[f"{statline}"].append(min(efficiencies_product_stats_values))
            stats_cells[f"{statline + 1}"].append(
                max(efficiencies_product_stats_values)
            )
            stats_cells[f"{statline + 2}"].append(
                mean(efficiencies_product_stats_values)
            )
            stats_cells[f"{statline + 3}"].append(
                median(efficiencies_product_stats_values)
            )
            statline += 4
        elif has_any_gamma and has_any_kappa:
            for line_letter in (f"{statline + j}" for j in range(4)):
                stats_cells[line_letter].append(_get_empty_cell())
            statline += 4

        if current_dataset_i in comparisons:
            dataset_comparison_stats, dataset_unique_reacs = comparisons[
                current_dataset_i
            ]
            tempstatline = statline
            for j, comparison_value in enumerate(dataset_comparison_stats.values()):
                stats_cells[f"{statline + j}"].append(comparison_value)
                tempstatline = statline + j
            statline = tempstatline + 1
            stats_cells[f"{statline}"].append(
                str(list(dataset_unique_reacs.values())[0])
            )
            stats_cells[f"{statline + 1}"].append(
                str(list(dataset_unique_reacs.values())[1])
            )
            statline += 2
        else:
            for line_letter in (f"{statline + j}" for j in range(8)):
                stats_cells[line_letter].append("(is best)")
            statline += 8

        if opt_dataset.with_kinetic_differences:
            unoptimized_reactions = get_unoptimized_reactions_in_nlp_solution(
                cobrak_model,
                opt_dataset.data,
                regard_iota=has_any_iota,
                regard_alpha=has_any_alpha,
            )
            prot_pool_sum = 0.0
            for reac_id, reac_data in cobrak_model.reactions.items():
                if reac_id not in opt_dataset.data:
                    continue
                if opt_dataset.data[reac_id] < min_var_value:
                    continue
                if reac_data.enzyme_reaction_data is None:
                    continue
                enzyme_var_id = get_reaction_enzyme_var_id(reac_id, reac_data)
                if enzyme_var_id not in opt_dataset.data:
                    continue
                enzyme_conc = opt_dataset.data[enzyme_var_id]
                mw = get_full_enzyme_mw(cobrak_model, reac_data)
                if reac_id in unoptimized_reactions:
                    ratio = (
                        unoptimized_reactions[reac_id][0]
                        / unoptimized_reactions[reac_id][1]
                    )
                    if ratio < 1.0:
                        ratio = 1.0
                    prot_pool_sum += mw * enzyme_conc * (ratio)
                else:
                    prot_pool_sum += mw * enzyme_conc

            stats_cells[f"{statline}"].append(prot_pool_sum)
            statline += 1
        elif has_any_kinetic_differences:
            stats_cells[f"{statline}"].append(" ")
            statline += 1

    for extrai, extravalues in enumerate(extra_optstatistics_data.values()):
        stats_cells[f"{statline + extrai}"].extend(
            [SpreadsheetCell(extravalue) for extravalue in extravalues]
        )

    # Reaction sheet
    reac_titles: list[Title] = [
        Title("ID", WIDTH_DEFAULT),
        Title("String", WIDTH_DEFAULT),
        Title("ΔG'° [kJ⋅mol⁻¹]", WIDTH_DEFAULT),
        Title("Enzyme(s)", WIDTH_DEFAULT),
        Title("kcat [h⁻¹]", WIDTH_DEFAULT),
    ]
    if show_regulation_coefficients:
        reac_titles += [
            Title("kms [M]", WIDTH_DEFAULT),
            Title("kis [M]", WIDTH_DEFAULT),
            Title("kas [M]", WIDTH_DEFAULT),
            Title("Hill coefficients [-]", WIDTH_DEFAULT),
        ]
    reac_cells: dict[str, list[str | float | int | bool | None | SpreadsheetCell]] = {
        reac_id: [] for reac_id in all_reac_ids
    }
    # Reaction data
    for reac_id in all_reac_ids:
        reaction = cobrak_model.reactions[reac_id]
        # Reac ID
        reac_cells[reac_id].append(reac_id)
        # Reac string
        reac_cells[reac_id].append(get_reaction_string(cobrak_model, reac_id))
        # Reac ΔG'°
        reac_cells[reac_id].append(str(_na_str_or_value(reaction.dG0)))
        enzyme_reaction_data = reaction.enzyme_reaction_data
        match enzyme_reaction_data:
            case None:
                enzyme_id = None
                k_cat = None
                k_ms = None
                k_is = None
                k_as = None
                hills = None
            case _:
                enzyme_id = str(enzyme_reaction_data.identifiers)
                k_cat = enzyme_reaction_data.k_cat
                k_ms = str(enzyme_reaction_data.k_ms)
                k_is = str(enzyme_reaction_data.k_is)
                k_as = str(enzyme_reaction_data.k_as)
                hills = str(enzyme_reaction_data.hill_coefficients)
        # Enzyme ID
        reac_cells[reac_id].append(enzyme_id)
        # kcat
        reac_cells[reac_id].append(k_cat)
        # kms
        reac_cells[reac_id].append(k_ms)
        if show_regulation_coefficients:
            # kis
            reac_cells[reac_id].append(k_is)
            # kas
            reac_cells[reac_id].append(k_as)
            # Hill coefficients
            reac_cells[reac_id].append(hills)

    # Variability data
    for var_dataset_name, var_dataset in variability_datasets.items():
        reac_titles.extend(
            (
                Title(var_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Min flux [mmol⋅gDW⁻¹⋅h⁻¹]", WIDTH_DEFAULT),
                Title("Max flux [mmol⋅gDW⁻¹⋅h⁻¹]", WIDTH_DEFAULT),
            )
        )
        if var_dataset.with_df:
            reac_titles.extend(
                (
                    Title("Min driving force [kJ⋅mol⁻¹]", WIDTH_DEFAULT),
                    Title("Max driving force [kJ⋅mol⁻¹]", WIDTH_DEFAULT),
                )
            )
        var_reac_ids = set(all_reac_ids) & set(var_dataset.data.keys())
        for reac_id in var_reac_ids:
            variability_tuple = var_dataset.data[reac_id]
            min_flux = variability_tuple[0]
            max_flux = variability_tuple[1]
            bg_color = _get_variability_bg_color(min_flux, max_flux)
            reac_cells[reac_id].append(
                SpreadsheetCell(min_flux, bg_color=bg_color, border=BORDER_BLACK_LEFT)
            )
            reac_cells[reac_id].append(SpreadsheetCell(max_flux, bg_color=bg_color))
            if var_dataset.with_df:
                df_var_id = f"{DF_VAR_PREFIX}{reac_id}"
                if df_var_id in var_dataset.data:
                    min_df = str(round(var_dataset.data[df_var_id][0], 4))
                    max_df = str(round(var_dataset.data[df_var_id][1], 4))
                else:
                    min_df = " "
                    max_df = " "
            else:
                min_df = " "
                max_df = " "
            reac_cells[reac_id].append(SpreadsheetCell(min_df, bg_color=bg_color))
            reac_cells[reac_id].append(SpreadsheetCell(max_df, bg_color=bg_color))
        missing_reac_ids = set(all_reac_ids) - set(var_dataset.data.keys())
        for missing_reac_id in missing_reac_ids:
            reac_cells[missing_reac_id].append(_get_empty_cell())
            reac_cells[missing_reac_id].append(_get_empty_cell())
            if var_dataset.with_df:
                reac_cells[missing_reac_id].append(_get_empty_cell())
                reac_cells[missing_reac_id].append(_get_empty_cell())

    # Optimization data
    for opt_dataset_name, opt_dataset in optimization_datasets.items():
        reac_titles.extend(
            (
                Title(opt_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Flux", WIDTH_DEFAULT),
            )
        )
        if opt_dataset.with_df:
            reac_titles.append(Title("Driving force [kJ⋅mol⁻¹]", WIDTH_DEFAULT))
        if opt_dataset.with_vplus:
            reac_titles.append(Title("V⁺ [mmol⋅gDW⁻¹⋅h⁻¹]", WIDTH_DEFAULT))
        if opt_dataset.with_kappa:
            reac_titles.append(Title("κ [0,1]", WIDTH_DEFAULT))
        if opt_dataset.with_gamma:
            reac_titles.append(Title("γ [0,1]", WIDTH_DEFAULT))
        if opt_dataset.with_iota:
            reac_titles.append(Title("ι [0,1]", WIDTH_DEFAULT))
        if opt_dataset.with_alpha:
            reac_titles.append(Title("α [0,1]", WIDTH_DEFAULT))
        if opt_dataset.with_kinetic_differences:
            reac_titles.append(Title('"Real" flux', WIDTH_DEFAULT))
            unoptimized_reactions = get_unoptimized_reactions_in_nlp_solution(
                cobrak_model,
                opt_dataset.data,
                regard_alpha=True,
                regard_iota=True,
            )
        opt_reac_ids = set(all_reac_ids) & set(opt_dataset.data.keys())
        reacs_with_too_low_flux = []
        for reac_id in opt_reac_ids:
            flux = get_fwd_rev_corrected_flux(
                reac_id=reac_id,
                usable_reac_ids=opt_reac_ids,
                result=opt_dataset.data,
                fwd_suffix=cobrak_model.fwd_suffix,
                rev_suffix=cobrak_model.rev_suffix,
            )
            if flux < min_var_value:
                reacs_with_too_low_flux.append(reac_id)
                continue
            bg_color = _get_optimization_bg_color(flux)
            reac_cells[reac_id].append(
                SpreadsheetCell(flux, bg_color=bg_color, border=BORDER_BLACK_LEFT)
            )
            enzyme_reaction_data = cobrak_model.reactions[reac_id].enzyme_reaction_data
            reaction = cobrak_model.reactions[reac_id]
            if opt_dataset.with_df:
                df_var_id = f"{DF_VAR_PREFIX}{reac_id}"
                if df_var_id in opt_dataset.data:
                    df_value = str(round(opt_dataset.data[df_var_id], 4))
                else:
                    df_value = " "
                reac_cells[reac_id].append(SpreadsheetCell(df_value, bg_color=bg_color))
            if opt_dataset.with_vplus:
                if enzyme_reaction_data is not None:
                    vplus = str(
                        enzyme_reaction_data.k_cat
                        * opt_dataset.data[
                            get_reaction_enzyme_var_id(reac_id, reaction)
                        ]
                    )
                else:
                    vplus = " "
                reac_cells[reac_id].append(SpreadsheetCell(vplus, bg_color=bg_color))
            if opt_dataset.with_kappa:
                kappa_var_id = KAPPA_VAR_PREFIX + reac_id
                if kappa_var_id in opt_dataset.data:
                    kappa_value = str(round(opt_dataset.data[kappa_var_id], 4))
                else:
                    kappa_value = " "
                reac_cells[reac_id].append(
                    SpreadsheetCell(kappa_value, bg_color=bg_color)
                )
            if opt_dataset.with_gamma:
                gamma_var_id = GAMMA_VAR_PREFIX + reac_id
                if gamma_var_id in opt_dataset.data:
                    gamma_value = str(round(opt_dataset.data[gamma_var_id], 4))
                else:
                    gamma_value = " "
                reac_cells[reac_id].append(
                    SpreadsheetCell(gamma_value, bg_color=bg_color)
                )
            if opt_dataset.with_iota:
                iota_var_id = IOTA_VAR_PREFIX + reac_id
                if iota_var_id in opt_dataset.data:
                    iota_value = str(opt_dataset.data[iota_var_id])
                else:
                    iota_value = " "
                reac_cells[reac_id].append(
                    SpreadsheetCell(iota_value, bg_color=bg_color)
                )
            if opt_dataset.with_alpha:
                alpha_var_id = ALPHA_VAR_PREFIX + reac_id
                if alpha_var_id in opt_dataset.data:
                    alpha_value = str(opt_dataset.data[alpha_var_id])
                else:
                    alpha_value = " "
                reac_cells[reac_id].append(
                    SpreadsheetCell(alpha_value, bg_color=bg_color)
                )
            if opt_dataset.with_kinetic_differences:
                if reac_id in unoptimized_reactions and (
                    round(
                        unoptimized_reactions[reac_id][1], kinetic_difference_precision
                    )
                    != round(
                        unoptimized_reactions[reac_id][0], kinetic_difference_precision
                    )
                ):
                    reac_cells[reac_id].append(
                        SpreadsheetCell(
                            unoptimized_reactions[reac_id][1], bg_color=bg_color
                        )
                    )
                else:
                    reac_cells[reac_id].append(
                        SpreadsheetCell(flux, bg_color=bg_color, font=FONT_ITALIC)
                    )
        missing_reac_ids = set(all_reac_ids) - set(opt_dataset.data.keys())
        missing_reac_ids |= set(reacs_with_too_low_flux)
        for missing_reac_id in missing_reac_ids:
            reac_cells[missing_reac_id].append(_get_empty_cell())
            num_extra = sum(
                [
                    opt_dataset.with_df,
                    opt_dataset.with_vplus,
                    opt_dataset.with_kappa,
                    opt_dataset.with_gamma,
                    opt_dataset.with_iota,
                    opt_dataset.with_alpha,
                    opt_dataset.with_kinetic_differences,
                ]
            )
            for _ in range(num_extra):
                reac_cells[missing_reac_id].append(_get_empty_cell())

    # Single enzyme sheet
    enzyme_titles: list[Title] = [
        Title("ID", WIDTH_DEFAULT),
        Title("MW", WIDTH_DEFAULT),
        Title("Conc. range [mmol⋅gDW⁻¹]", WIDTH_DEFAULT),
    ]
    enzyme_cells: dict[str, list[str | float | int | bool | None | SpreadsheetCell]] = {
        enzyme_id: [] for enzyme_id in all_enzyme_ids
    }
    # Single enzyme data
    for enzyme_id in all_enzyme_ids:
        enzyme: Enzyme = cobrak_model.enzymes[enzyme_id]
        # Enzyme ID
        enzyme_cells[enzyme_id].append(enzyme_id)
        # Enzyme MW
        enzyme_cells[enzyme_id].append(enzyme.molecular_weight)
        # Enzyme concentration range
        match enzyme.min_conc:
            case None:
                min_conc = None
                bg_color = BG_COLOR_BLACK
            case _:
                min_conc = enzyme.min_conc
                bg_color = BG_COLOR_DEFAULT
        match enzyme.max_conc:
            case None:
                max_conc = None
                bg_color = BG_COLOR_BLACK
            case _:
                max_conc = enzyme.max_conc
                bg_color = BG_COLOR_DEFAULT

        enzyme_cells[enzyme_id].append(SpreadsheetCell(min_conc, bg_color=bg_color))
        enzyme_cells[enzyme_id].append(SpreadsheetCell(min_conc, bg_color=bg_color))

    # Enzyme complexes sheet
    enzcomplex_titles: list[Title] = [
        Title("ID", WIDTH_DEFAULT),
        Title("Reactions", WIDTH_DEFAULT),
        Title("MW", WIDTH_DEFAULT),
    ]
    enzcomplex_cells: dict[
        str, list[str | float | int | bool | None | SpreadsheetCell]
    ] = {enzcomplex_id: [] for enzcomplex_id in all_enzcomplex_ids}
    # Enzyme complex data
    for enzcomplex_id in all_enzcomplex_ids:
        reac_id, reaction = _get_enzcomplex_reaction(cobrak_model, enzcomplex_id)
        # Enzyme complex ID
        enzcomplex_cells[enzcomplex_id].append(
            enzcomplex_id.replace(ENZYME_VAR_PREFIX, "").split(ENZYME_VAR_INFIX)[0]
        )
        # Associated reaction
        if reaction.enzyme_reaction_data is None:
            raise ValueError
        if reaction.enzyme_reaction_data.identifiers == [""]:
            continue
        enzcomplex_cells[enzcomplex_id].append(reac_id)
        # Enzyme complex MW
        full_mw = get_full_enzyme_mw(cobrak_model, reaction)
        enzcomplex_cells[enzcomplex_id].append(full_mw)

    # Variability data
    for var_dataset_name, var_dataset in variability_datasets.items():
        enzcomplex_titles.extend(
            (
                Title(var_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Min conc. [mmol⋅gDW⁻¹]", WIDTH_DEFAULT),
                Title("Max conc. [mmolgDW⁻¹]", WIDTH_DEFAULT),
            )
        )
        var_enzcomplex_ids = set(all_enzcomplex_ids) & set(var_dataset.data.keys())
        for enzcomplex_id in var_enzcomplex_ids:
            _, reaction = _get_enzcomplex_reaction(cobrak_model, enzcomplex_id)
            variability_tuple = var_dataset.data[enzcomplex_id]
            min_conc = variability_tuple[0]
            max_conc = variability_tuple[1]
            bg_color = _get_variability_bg_color(min_conc, max_conc)
            enzcomplex_cells[enzcomplex_id].append(
                SpreadsheetCell(min_conc, bg_color=bg_color, border=BORDER_BLACK_LEFT)
            )
            enzcomplex_cells[enzcomplex_id].append(
                SpreadsheetCell(max_conc, bg_color=bg_color)
            )
        missing_enzcomplex_ids = set(all_enzcomplex_ids) - set(var_dataset.data.keys())
        for missing_enzcomplex_id in missing_enzcomplex_ids:
            enzcomplex_cells[missing_enzcomplex_id].append(_get_empty_cell())
            enzcomplex_cells[missing_enzcomplex_id].append(_get_empty_cell())

    # Enzyme complex data
    for opt_dataset_name, opt_dataset in optimization_datasets.items():
        enzcomplex_titles.extend(
            (
                Title(opt_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Concentration [mmol⋅gDW⁻¹]", WIDTH_DEFAULT),
                Title("% of pool", WIDTH_DEFAULT),
            )
        )
        opt_enzcomplex_ids = set(all_enzcomplex_ids) & set(opt_dataset.data.keys())
        for enzcomplex_id in opt_enzcomplex_ids:
            _, reaction = _get_enzcomplex_reaction(cobrak_model, enzcomplex_id)
            complexconc = opt_dataset.data[enzcomplex_id]
            pool_pct = (
                100
                * complexconc
                * get_full_enzyme_mw(cobrak_model, reaction)
                / cobrak_model.max_prot_pool
            )
            bg_color = _get_optimization_bg_color(complexconc)
            enzcomplex_cells[enzcomplex_id].append(
                SpreadsheetCell(
                    complexconc, bg_color=bg_color, border=BORDER_BLACK_LEFT
                )
            )
            enzcomplex_cells[enzcomplex_id].append(
                SpreadsheetCell(round(pool_pct, 4), bg_color=bg_color)
            )
        missing_enzcomplex_ids = set(all_enzcomplex_ids) - set(opt_dataset.data.keys())
        for missing_enzcomplex_id in missing_enzcomplex_ids:
            enzcomplex_cells[missing_enzcomplex_id].append(_get_empty_cell())
            enzcomplex_cells[missing_enzcomplex_id].append(_get_empty_cell())

    # Metabolite sheet
    met_titles: list[Title] = [
        Title("ID", WIDTH_DEFAULT),
        Title("Min set concentration [mmol⋅gDW⁻¹⋅h⁻¹)]", WIDTH_DEFAULT),
        Title("Max set concentration [mmolgDW⁻¹⋅h⁻¹)]", WIDTH_DEFAULT),
        Title("Annotation", WIDTH_DEFAULT),
    ]
    met_cells: dict[str, list[str | float | int | bool | None | SpreadsheetCell]] = {
        met_id: [] for met_id in all_met_ids
    }
    # Metabolite data
    for met_id in all_met_ids:
        met: Metabolite = cobrak_model.metabolites[met_id]
        # Met ID
        met_cells[met_id].append(met_id)
        # Min conc
        met_cells[met_id].append(exp(met.log_min_conc))
        # Max conc
        met_cells[met_id].append(exp(met.log_max_conc))
        # Annotation
        met_cells[met_id].append(str(met.annotation))

    # Variability data
    for var_dataset_name, var_dataset in variability_datasets.items():
        met_titles.extend(
            (
                Title(var_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Min concentration [mmol⋅gDW⁻¹⋅h⁻¹)]", WIDTH_DEFAULT),
                Title("Max concentration [mmol⋅gDW⁻¹⋅h⁻¹)]", WIDTH_DEFAULT),
            )
        )
        all_met_var_ids = [LNCONC_VAR_PREFIX + met_id for met_id in all_met_ids]
        var_met_ids = set(all_met_var_ids) & set(var_dataset.data.keys())
        for met_var_id in var_met_ids:
            variability_tuple = var_dataset.data[met_var_id]
            min_conc = exp(variability_tuple[0])
            max_conc = exp(variability_tuple[1])
            bg_color = BG_COLOR_RED if min_conc == max_conc else BG_COLOR_GREEN
            met_cells[_get_met_id_from_met_var_id(met_var_id)].append(
                SpreadsheetCell(min_conc, bg_color=bg_color, border=BORDER_BLACK_LEFT)
            )
            met_cells[_get_met_id_from_met_var_id(met_var_id)].append(
                SpreadsheetCell(max_conc, bg_color=bg_color)
            )
        missing_met_var_ids = set(all_met_var_ids) - set(var_dataset.data.keys())
        for missing_met_var_id in missing_met_var_ids:
            met_cells[_get_met_id_from_met_var_id(missing_met_var_id)].append(
                _get_empty_cell()
            )
            met_cells[_get_met_id_from_met_var_id(missing_met_var_id)].append(
                _get_empty_cell()
            )

    # Optimization data
    for opt_dataset_name, opt_dataset in optimization_datasets.items():
        met_titles.extend(
            (
                Title(opt_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Concentration [M]", WIDTH_DEFAULT),
                Title("Consumption [mmol⋅gDW⁻¹⋅h⁻¹]", WIDTH_DEFAULT),
                Title("Production [mmol⋅gDW⁻¹⋅h⁻¹]", WIDTH_DEFAULT),
            )
        )
        opt_met_ids = set(all_met_var_ids) & set(opt_dataset.data.keys())
        for met_var_id in opt_met_ids:
            conc = exp(opt_dataset.data[met_var_id])
            consumption, production = get_metabolite_consumption_and_production(
                cobrak_model, _get_met_id_from_met_var_id(met_var_id), opt_dataset.data
            )
            bg_color = _get_optimization_bg_color(consumption)
            met_cells[_get_met_id_from_met_var_id(met_var_id)].append(
                SpreadsheetCell(conc, bg_color=bg_color, border=BORDER_BLACK_LEFT)
            )
            met_cells[_get_met_id_from_met_var_id(met_var_id)].append(
                SpreadsheetCell(consumption, bg_color=bg_color)
            )
            met_cells[_get_met_id_from_met_var_id(met_var_id)].append(
                SpreadsheetCell(production, bg_color=bg_color)
            )
        missing_met_ids = set(all_met_var_ids) - set(opt_dataset.data.keys())
        for missing_met_id in missing_met_ids:
            for _ in range(3):
                met_cells[_get_met_id_from_met_var_id(missing_met_id)].append(
                    _get_empty_cell()
                )

    # κ and γ statistics
    kgstats_titles: list[Title] = [Title("Rank", WIDTH_DEFAULT, is_metatitle=False)]
    kgstats_cells: dict[
        str, list[str | float | int | bool | None | SpreadsheetCell]
    ] = {str(i): [i + 1] for i in range(len(cobrak_model.reactions))}
    for opt_dataset_name, opt_dataset in optimization_datasets.items():
        kgstats_titles.extend(
            (
                Title(opt_dataset_name, WIDTH_DEFAULT, is_metatitle=True),
                Title("Reaction ID", WIDTH_DEFAULT),
                Title("κ", WIDTH_DEFAULT),
                Title("Reaction ID", WIDTH_DEFAULT),
                Title("γ", WIDTH_DEFAULT),
                Title("Reaction ID", WIDTH_DEFAULT),
                Title("ι", WIDTH_DEFAULT),
                Title("Reaction ID", WIDTH_DEFAULT),
                Title("α", WIDTH_DEFAULT),
                Title("Reaction ID", WIDTH_DEFAULT),
                Title(kappa_gamma_iota_alpha_str, WIDTH_DEFAULT),
            )
        )

        _, kappa_stats, gamma_stats, iota_stats, alpha_stats, multiplier_stats = (
            get_df_and_efficiency_factors_sorted_lists(
                cobrak_model,
                opt_dataset.data,
                min_var_value,
            )
        )
        kappa_stats_titles = list(kappa_stats.keys())
        gamma_stats_titles = list(gamma_stats.keys())
        iota_stats_titles = list(iota_stats.keys())
        alpha_stats_titles = list(alpha_stats.keys())
        kappa_times_gamma_stats_titles = list(multiplier_stats.keys())
        for key, cell_list in kgstats_cells.items():
            # κ
            if len(kappa_stats_titles) > int(key):
                cell_list.extend(
                    (
                        kappa_stats_titles[int(key)],
                        kappa_stats[kappa_stats_titles[int(key)]],
                    )
                )
            else:
                cell_list.extend((None, None))
            # γ
            if len(gamma_stats_titles) > int(key):
                cell_list.extend(
                    (
                        gamma_stats_titles[int(key)],
                        gamma_stats[gamma_stats_titles[int(key)]],
                    )
                )
            else:
                cell_list.extend((None, None))
            # ι
            if len(iota_stats_titles) > int(key):
                cell_list.extend(
                    (
                        iota_stats_titles[int(key)],
                        iota_stats[iota_stats_titles[int(key)]],
                    )
                )
            else:
                cell_list.extend((None, None))
            # α
            if len(alpha_stats_titles) > int(key):
                cell_list.extend(
                    (
                        alpha_stats_titles[int(key)],
                        alpha_stats[alpha_stats_titles[int(key)]],
                    )
                )
            else:
                cell_list.extend((None, None))
            # κ⋅γ⋅ι⋅α
            if len(kappa_times_gamma_stats_titles) > int(key):
                cell_list.extend(
                    (
                        kappa_times_gamma_stats_titles[int(key)],
                        multiplier_stats[kappa_times_gamma_stats_titles[int(key)]][0],
                    )
                )
            else:
                cell_list.extend((None, None))

    titles_and_data_dict: dict[
        str,
        tuple[
            list[Title],
            dict[str, list[str | float | int | bool | None | SpreadsheetCell]],
        ],
    ] = {
        "Index": (index_titles, index_cells),
        "A) Optimization statistics": (stats_titles, stats_cells),
        "B) Model settings": (model_titles, model_cells),
        "C) Reactions": (reac_titles, reac_cells),
        "D) Metabolites": (met_titles, met_cells),
        "E) Enzymes": (enzyme_titles, enzyme_cells),
        "F) Complexes": (enzcomplex_titles, enzcomplex_cells),
    }
    if has_any_gamma or has_any_kappa:
        titles_and_data_dict |= {
            "G) Efficiency factor statistics": (kgstats_titles, kgstats_cells),
        }

    # Correction data (if given)
    correction_titles: list[Title] = [
        Title("Affected parameter", WIDTH_DEFAULT),
        Title("Original value", WIDTH_DEFAULT),
    ]
    correction_cells: dict[
        str, list[str | float | int | bool | None | SpreadsheetCell]
    ] = {}

    num_processed_datasets = 0
    for opt_dataset_name, opt_dataset in optimization_datasets.items():
        if not opt_dataset.with_error_corrections:
            num_processed_datasets += 1
            continue
        correction_titles.append(Title(opt_dataset_name, WIDTH_DEFAULT))

        for var_name, var_value in opt_dataset.data.items():
            displayed_var = var_name.replace(ERROR_VAR_PREFIX + "_", "")
            if not var_name.startswith(ERROR_VAR_PREFIX):
                if displayed_var in correction_cells:
                    correction_cells[displayed_var].append(None)
                continue

            round_value = 12
            is_dataset_dependent: bool = False
            is_relative: bool = True
            original_value: float = 0.0
            if "kcat_times_e_" in displayed_var:
                reac_id = displayed_var.split("kcat_times_e_")[1]
                enzyme_id = get_reaction_enzyme_var_id(
                    reac_id, cobrak_model.reactions[reac_id]
                )
                original_value = (
                    cobrak_model.reactions[reac_id].enzyme_reaction_data.k_cat
                    * opt_dataset.data[enzyme_id]
                )
                displayed_original_value = "(see comments)"
                error_value = var_value - original_value
                is_dataset_dependent = True
            elif displayed_var.endswith(("_substrate", "_product")):
                reac_id = displayed_var.split("____")[0]
                met_id = (
                    displayed_var.split("____")[1]
                    .replace("_substrate", "")
                    .replace("_product", "")
                )
                original_value = cobrak_model.reactions[
                    reac_id
                ].enzyme_reaction_data.k_ms[met_id]
                displayed_original_value = original_value
                error_mult = +1 if displayed_var.endswith("_product") else -1
                error_value = error_mult * -(
                    original_value - exp(log(original_value) + error_mult * var_value)
                )
                round_value = 12
                if displayed_var.endswith("_substrate"):
                    print(
                        displayed_var,
                        original_value,
                        error_value,
                        abs(error_value) / original_value,
                    )
            elif displayed_var.startswith("dG0_"):
                reac_id = displayed_var[len("dG0_") :]
                original_value = cobrak_model.reactions[reac_id].dG0
                is_relative = False
            elif displayed_var.endswith(("_plus", "_minus")):
                valueblock = displayed_var.split("_origstart_")[1].split("_origend_")[0]
                min_value = float(valueblock.split("__")[0].replace("-", "."))
                max_value = float(valueblock.split("__")[1].replace("-", "."))
                displayed_original_value = f"({min_value}, {max_value}"
                min_difference = abs(var_value - min_value)
                max_difference = abs(var_value - max_value)
                original_value = (
                    min_value if min_difference < max_difference else max_value
                )
                error_value = min(max_difference, min_difference)
            else:
                continue

            if not is_relative:
                print(displayed_var, error_value, original_value)
            if not (
                is_relative and (error_value / original_value) >= min_rel_correction
            ) or (not is_relative and error_value >= min_var_value):
                if displayed_var in correction_cells:
                    correction_cells[displayed_var].append(None)
                continue

            if displayed_var not in correction_cells:
                correction_cells[displayed_var] = [
                    displayed_var,
                    displayed_original_value,
                ] + [None for _ in range(num_processed_datasets)]
            correction_cells[displayed_var].append(
                f"{round(error_value, round_value)}{f' from {round(original_value, round_value)}' if is_dataset_dependent else ''}"
            )

        num_processed_datasets += 1

    if correction_cells != {}:
        titles_and_data_dict[
            f"{'H' if has_any_alpha or has_any_iota or has_any_gamma or has_any_kappa else 'G'}) Corrections"
        ] = (correction_titles, correction_cells)

    _create_xlsx_from_datadicts(
        path=path,
        titles_and_data_dict=titles_and_data_dict,
    )
